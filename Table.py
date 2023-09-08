# Python module to handle excel tables.
import openpyxl

# This class is where the old table is dissected.
class Table:
    def __init__(self, directory:str):
        self.table = openpyxl.load_workbook(directory)
        self.activeSheet = self.table.active
        self.sheets = dict()
        for worksheet in self.table.worksheets:
            self.sheets.update({worksheet.title: worksheet})

    def setActiveSheet(self, newSheet):
        try:
            self.activeSheet = self.sheets[newSheet]
        except:
            print("[ERROR] Unknow Sheet - Check if the sheet exists in the Table.")

    def getActiveSheetElement(self, element):
        elementDic = {"columns": self.activeSheet.iter_cols(values_only=True),
                       "rows": self.activeSheet.iter_rows(values_only=True)}
        try:
            templist = list()
            for x in elementDic[element]:
                # Filter to remove None elements.
                x = list(filter(lambda y: False if y is None else True, x))
                templist.append(x)
            # Filter to remove Empty elements and the Header.
            templist.pop(0)
            templist = list(filter(lambda y: False if len(y) == 0 else True, templist))
        except:
            print("[ERROR] Unknow Element - Check if the passed element exists.")
        return templist